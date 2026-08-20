import io
import zipfile

from openpyxl import load_workbook


def _create_project(client, headers, name="Retail"):
    return client.post("/api/v1/projects", json={"name": name}, headers=headers).json()["id"]


def _create_entity_with_field(client, headers, project_id, entity_name, field_name="name"):
    entity = client.post(
        f"/api/v1/projects/{project_id}/entities", json={"name": entity_name}, headers=headers
    ).json()
    client.post(
        f"/api/v1/projects/{project_id}/entities/{entity['id']}/fields",
        json={
            "name": field_name,
            "field_type": "string",
            "required": True,
            "nullable": False,
        },
        headers=headers,
    )
    return entity["id"]


def test_entity_generate_xlsx(client, auth_headers):
    project_id = _create_project(client, auth_headers)
    entity_id = _create_entity_with_field(client, auth_headers, project_id, "Customer")
    base = f"/api/v1/projects/{project_id}/entities/{entity_id}"

    resp = client.post(f"{base}/generate?format=xlsx", json={"count": 5}, headers=auth_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("name",)
    assert len(rows) == 6  # header + 5 rows


def test_project_generate_xlsx_has_one_sheet_per_entity(client, auth_headers):
    project_id = _create_project(client, auth_headers)
    _create_entity_with_field(client, auth_headers, project_id, "Customer")
    _create_entity_with_field(client, auth_headers, project_id, "Order", field_name="ref")

    resp = client.post(
        f"/api/v1/projects/{project_id}/generate?format=xlsx",
        json={"counts": {}, "count": 4},
        headers=auth_headers,
    )
    assert resp.status_code == 200

    wb = load_workbook(io.BytesIO(resp.content))
    assert set(wb.sheetnames) == {"Customer", "Order"}
    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        assert len(rows) == 5  # header + 4 rows


def test_project_generate_csv_zip(client, auth_headers):
    project_id = _create_project(client, auth_headers)
    _create_entity_with_field(client, auth_headers, project_id, "Customer")
    _create_entity_with_field(client, auth_headers, project_id, "Order", field_name="ref")

    resp = client.post(
        f"/api/v1/projects/{project_id}/generate?format=csv",
        json={"count": 3},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/zip"

    zf = zipfile.ZipFile(io.BytesIO(resp.content))
    assert set(zf.namelist()) == {"Customer.csv", "Order.csv"}
    customer_csv = zf.read("Customer.csv").decode()
    lines = customer_csv.strip().splitlines()
    assert lines[0] == "name"
    assert len(lines) == 4  # header + 3 rows


def test_xlsx_includes_workflow_history_column(client, auth_headers):
    project_id = _create_project(client, auth_headers)
    entity = client.post(
        f"/api/v1/projects/{project_id}/entities", json={"name": "Order"}, headers=auth_headers
    ).json()
    field = client.post(
        f"/api/v1/projects/{project_id}/entities/{entity['id']}/fields",
        json={"name": "status", "field_type": "string", "required": True, "nullable": False},
        headers=auth_headers,
    ).json()
    client.post(
        f"/api/v1/projects/{project_id}/entities/{entity['id']}/workflows",
        json={
            "field_id": field["id"],
            "states": ["created", "shipped"],
            "initial_states": ["created"],
            "transitions": [{"source": "created", "target": "shipped"}],
        },
        headers=auth_headers,
    )

    base = f"/api/v1/projects/{project_id}/entities/{entity['id']}"
    resp = client.post(f"{base}/generate?format=xlsx", json={"count": 3}, headers=auth_headers)
    assert resp.status_code == 200

    wb = load_workbook(io.BytesIO(resp.content))
    header = next(wb.active.iter_rows(values_only=True))
    assert header == ("status", "status_history")
