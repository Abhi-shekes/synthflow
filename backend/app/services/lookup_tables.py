"""Parses an uploaded lookup-table file (CSV, Excel, or JSON) into a
column list and a list of row dicts — see app.models.lookup_table for how
the result is stored and app.services.generator.build_lookup_pools for how
a field draws from it.
"""

import csv
import io
import json
from typing import Any

from openpyxl import load_workbook


class LookupParseError(ValueError):
    pass


def coerce_numeric(value: Any) -> Any:
    """Text that looks numeric becomes a real int/float; anything else is
    left untouched. CSV and Excel cells come back as text even when they're
    numbers, unlike JSON which already carries native types — this is only
    called where that text-vs-native gap actually exists: here for parsed
    CSV cells, and reused by app.services.generator for weighted-enum
    values, which are always configured as strings (app.models.field's
    `enum_values: list[str]`) but should come out as real numbers in
    generated output when they look like one, e.g. an HTTP status code
    enum ("200", "404", "500") generating real ints instead of strings."""
    if value is None or value == "":
        return None
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        pass
    return value


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [{k: coerce_numeric(v) for k, v in row.items() if k is not None} for row in reader]


def _parse_excel(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    columns = [str(c) if c is not None else "" for c in header]
    return [dict(zip(columns, row, strict=False)) for row in rows_iter]


def _parse_json(content: bytes) -> list[dict[str, Any]]:
    try:
        data = json.loads(content)
    except json.JSONDecodeError as exc:
        raise LookupParseError(f"Invalid JSON: {exc}") from exc
    if not isinstance(data, list) or not all(isinstance(row, dict) for row in data):
        raise LookupParseError("JSON lookup table must be a list of flat objects")
    return data


def parse_upload(
    filename: str, content: bytes, max_rows: int
) -> tuple[list[str], list[dict[str, Any]]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        rows = _parse_csv(content)
    elif lower.endswith((".xlsx", ".xls")):
        rows = _parse_excel(content)
    elif lower.endswith(".json"):
        rows = _parse_json(content)
    else:
        raise LookupParseError("Unsupported file type — use .csv, .xlsx, or .json")

    if not rows:
        raise LookupParseError("File contains no rows")
    if len(rows) > max_rows:
        raise LookupParseError(f"File has {len(rows)} rows, which exceeds the {max_rows}-row limit")

    columns: list[str] = []
    seen = set()
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                columns.append(key)

    return columns, rows
