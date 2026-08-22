"""Batch generation engine: turns Entity field definitions into fake rows.

Phase 1 covers a single entity in isolation. Phase 2 adds `generate_project`
(entities generated in relationship-dependency order so a child's foreign-key
field draws real values from its already-generated parent), formula fields
(a field's value computed from other fields on the same row instead of being
randomized), rules (a boolean expression a generated row must satisfy,
enforced by discard-and-retry), and workflows (a field's value comes from a
random walk over a state machine instead of being randomized independently).
Phase 4 adds trends (a numeric field's value as a function of its row's
position within the current batch — see app.models.trend.Trend) and weighted
enum fields (app.models.field.EntityField.enum_weights).
"""

import csv
import io
import json
import random
import re
import uuid
import zipfile
from collections.abc import Iterable, Iterator
from datetime import UTC, datetime, timedelta
from typing import Any

import exrex
from faker import Faker
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models.entity import Entity
from app.models.error_injection import ErrorInjection, ErrorType
from app.models.event_trigger import EventTrigger
from app.models.field import EntityField, FieldType
from app.models.geo_route import GeoRoute
from app.models.lookup_attachment import LookupAttachment
from app.models.relationship import Relationship, RelationshipType
from app.models.rule import Rule
from app.models.trend import Trend
from app.models.workflow import Workflow
from app.services.expressions import ExpressionError, evaluate
from app.services.geo_routes import generate_geo_point
from app.services.lookup_tables import coerce_numeric
from app.services.plugins import generate_preset_value
from app.services.quality.diagnostics import GenerationDiagnostics
from app.services.trends import generate_trend_value

faker = Faker()

NULLABLE_PROBABILITY = 0.15
MAX_UNIQUE_ATTEMPTS = 100
MAX_RULE_ATTEMPTS = 200
MAX_WORKFLOW_STEPS = 20
WORKFLOW_STOP_PROBABILITY = 0.35


def _generate_value(field: EntityField) -> Any:
    if field.field_type == FieldType.STRING:
        if field.preset:
            return generate_preset_value(field.preset)
        if field.regex:
            return exrex.getone(field.regex)
        return faker.word()

    if field.field_type == FieldType.INTEGER:
        low = int(field.min_value) if field.min_value is not None else 0
        high = int(field.max_value) if field.max_value is not None else 100_000
        return random.randint(low, high)

    if field.field_type == FieldType.FLOAT:
        low = field.min_value if field.min_value is not None else 0.0
        high = field.max_value if field.max_value is not None else 100_000.0
        return round(random.uniform(low, high), 2)

    if field.field_type == FieldType.BOOLEAN:
        return faker.boolean()

    if field.field_type == FieldType.DATE:
        start = datetime.now(UTC).date() - timedelta(days=365)
        return faker.date_between(start_date=start, end_date="today").isoformat()

    if field.field_type == FieldType.DATETIME:
        start = datetime.now(UTC) - timedelta(days=365)
        return faker.date_time_between(start_date=start, end_date="now").isoformat()

    if field.field_type == FieldType.UUID:
        return str(uuid.uuid4())

    if field.field_type == FieldType.ENUM:
        if not field.enum_values:
            raise ValueError(f"Field '{field.name}' is type enum but has no enum_values")
        if field.enum_weights:
            chosen = random.choices(field.enum_values, weights=field.enum_weights, k=1)[0]
        else:
            chosen = random.choice(field.enum_values)
        # enum_values are always configured as strings (see EntityField), but
        # a numeric-looking one — e.g. a weighted HTTP status code enum
        # ("200", "404", "500") for API-behavior simulation — should come out
        # as a real int/float in generated output, not stay a string.
        return coerce_numeric(chosen)

    if field.field_type == FieldType.ARRAY:
        return [faker.word() for _ in range(random.randint(1, 3))]

    if field.field_type in (FieldType.OBJECT, FieldType.JSON):
        return {faker.word(): faker.word() for _ in range(random.randint(1, 3))}

    raise ValueError(f"Unsupported field type: {field.field_type}")


def _generate_unique_value(
    field: EntityField, seen: set, diagnostics: GenerationDiagnostics | None = None
) -> Any:
    for attempt in range(MAX_UNIQUE_ATTEMPTS):
        value = _generate_value(field)
        key = str(value)
        if key not in seen:
            seen.add(key)
            # `attempt` is how many *collisions* preceded this value. Rising
            # towards MAX_UNIQUE_ATTEMPTS is the only warning available that
            # the pool is nearly exhausted before it fails outright.
            if diagnostics is not None:
                diagnostics.unique_retry(field.name, attempt)
            return value
    raise ValueError(
        f"Could not generate a unique value for field '{field.name}' "
        f"after {MAX_UNIQUE_ATTEMPTS} attempts"
    )


def build_lookup_pools(attachments: list[LookupAttachment]) -> dict[str, list[Any]]:
    """Builds a field-name -> value-pool dict from each attachment's lookup
    table column, in the exact shape `generate_rows` already expects for
    relationship-sourced `fk_pools` — a lookup-attached field is drawn from
    this pool the same way a foreign-key field is, including honoring
    `field.unique` for without-replacement draws (see `_generate_one_row`).
    Rows missing the column, or holding a null there, are skipped."""
    pools: dict[str, list[Any]] = {}
    for attachment in attachments:
        values = [
            row[attachment.column]
            for row in attachment.lookup_table.data
            if row.get(attachment.column) is not None
        ]
        if not values:
            raise ValueError(
                f"Lookup table '{attachment.lookup_table.name}' has no non-null values "
                f"in column '{attachment.column}'"
            )
        pools[attachment.field.name] = values
    return pools


def _evaluate_formula(field: EntityField, row_so_far: dict[str, Any]) -> Any:
    try:
        return evaluate(field.formula, row_so_far)
    except ExpressionError as exc:
        raise ValueError(f"Formula for field '{field.name}' failed: {exc}") from exc


def _first_failing_rule(
    row: dict[str, Any], rules: list[Rule], cross_entity_context: dict[str, dict[str, Any]]
) -> Rule | None:
    """The first rule this row fails, or None if it satisfies all of them.

    Returns the rule rather than a bool so a discarded candidate can be
    attributed to what rejected it — "rule X threw away 95% of candidates"
    is actionable in a way that "5% of rows survived" is not. Stopping at
    the first failure keeps the counts summing to the discard total instead
    of double-counting a candidate that several rules would have rejected.
    """
    variables = {**row, **cross_entity_context} if cross_entity_context else row
    for rule in rules:
        try:
            if not evaluate(rule.condition, variables):
                return rule
        except ExpressionError as exc:
            raise ValueError(f"Rule '{rule.condition}' failed to evaluate: {exc}") from exc
    return None


def _evaluate_event_triggers(
    row: dict[str, Any],
    event_triggers: list[EventTrigger],
    cross_entity_context: dict[str, dict[str, Any]],
) -> list[str]:
    """Unlike a rule, a matching trigger doesn't reject the row — it collects
    labels for `_triggered_events`. See EventTrigger's docstring for why
    that's the whole feature for now (no external notification fires)."""
    variables = {**row, **cross_entity_context} if cross_entity_context else row
    triggered: list[str] = []
    for trigger in event_triggers:
        try:
            if evaluate(trigger.condition, variables):
                triggered.append(trigger.label)
        except ExpressionError as exc:
            raise ValueError(f"Event trigger '{trigger.label}' failed to evaluate: {exc}") from exc
    return triggered


def _generate_state_walk(workflow: Workflow) -> list[str]:
    """A random walk through the workflow's transition graph, starting from a
    random initial state and stopping (with WORKFLOW_STOP_PROBABILITY chance
    per step by default, or a per-state override from `stop_probabilities`,
    or when a state has no outgoing transitions) within MAX_WORKFLOW_STEPS
    hops. Later states are naturally rarer since they require surviving more
    consecutive "don't stop" draws — a deliberate, simple stand-in for "most
    records are further along than not yet started, but few reach the very
    end," not a claim about any real-world process. A state with multiple
    outgoing transitions picks among them by `weight` (default 1.0, i.e.
    uniform) rather than always uniformly — together with per-state stop
    probabilities, this is what makes a linear chain a realistic funnel with
    asymmetric drop-off per stage, not just a flat one."""
    if not workflow.initial_states:
        raise ValueError("Workflow has no initial states")

    by_source: dict[str, list[tuple[str, float]]] = {}
    for t in workflow.transitions:
        by_source.setdefault(t["source"], []).append((t["target"], t.get("weight", 1.0)))
    stop_probabilities = workflow.stop_probabilities or {}

    path = [random.choice(workflow.initial_states)]
    for _ in range(MAX_WORKFLOW_STEPS - 1):
        options = by_source.get(path[-1], [])
        stop_probability = stop_probabilities.get(path[-1], WORKFLOW_STOP_PROBABILITY)
        if not options or random.random() < stop_probability:
            break
        targets, weights = zip(*options, strict=True)
        path.append(random.choices(targets, weights=weights, k=1)[0])
    return path


def advance_state(workflow: Workflow, current: str) -> str:
    """One step onward from `current`, for a record seen again later.

    `_generate_state_walk` is right for a *batch*: each row gets its own
    fresh walk, so a batch catches records at different points in a process,
    which is what makes a funnel look like a funnel. It is wrong for the
    *same* record seen twice — a customer who reached "checkout" yesterday
    has not gone back to "signed up" today. This is the Phase 13 half of
    that: given where a record actually is, take at most one step.

    The same weights and per-state stop probabilities apply, so a record's
    progress over many updates traces the same distribution a single walk
    would have produced in one go. A terminal state, or a roll that stops,
    returns `current` unchanged — records that have finished stay finished
    rather than being pushed somewhere they cannot go.

    An unknown `current` (the workflow was edited under a stored record)
    restarts from an initial state rather than raising: the alternative is a
    generation call that fails permanently because of a schema edit nobody
    connects to it.
    """
    if current not in workflow.states:
        return random.choice(workflow.initial_states)

    by_source: dict[str, list[tuple[str, float]]] = {}
    for t in workflow.transitions:
        by_source.setdefault(t["source"], []).append((t["target"], t.get("weight", 1.0)))

    options = by_source.get(current, [])
    stop_probability = (workflow.stop_probabilities or {}).get(current, WORKFLOW_STOP_PROBABILITY)
    if not options or random.random() < stop_probability:
        return current
    targets, weights = zip(*options, strict=True)
    return random.choices(targets, weights=weights, k=1)[0]


def _corrupt_value(
    value: Any,
    field: EntityField,
    error_types: list[str],
    previous_row: dict[str, Any] | None,
) -> Any:
    """Replaces an already-computed field value with a deliberately bad one.
    Picks one enabled error type at random per corrupted row, so a field with
    several configured types produces a mix of failure modes across a batch
    rather than always the same one. See app.models.error_injection for the
    full design, including the documented rules interaction."""
    error_type = random.choice(error_types)

    if error_type == ErrorType.NULL:
        return None

    if error_type == ErrorType.EMPTY:
        if field.field_type == FieldType.STRING:
            return ""
        if field.field_type == FieldType.ARRAY:
            return []
        return {}

    if error_type == ErrorType.DUPLICATE:
        if previous_row is not None and field.name in previous_row:
            return previous_row[field.name]
        return value

    if error_type == ErrorType.TRUNCATE:
        text = str(value)
        if len(text) <= 1:
            return text
        return text[: random.randint(1, len(text) - 1)]

    if error_type == ErrorType.WRONG_TYPE:
        if field.field_type in (FieldType.INTEGER, FieldType.FLOAT):
            return faker.word()
        return faker.random_int(min=0, max=999_999)

    if error_type == ErrorType.OUT_OF_RANGE:
        if field.field_type == FieldType.INTEGER:
            low = int(field.min_value) if field.min_value is not None else 0
            high = int(field.max_value) if field.max_value is not None else 100_000
            offset = random.randint(1, 1_000)
            return random.choice([low - offset, high + offset])
        low = field.min_value if field.min_value is not None else 0.0
        high = field.max_value if field.max_value is not None else 100_000.0
        offset = random.uniform(1, 1_000)
        return random.choice([low - offset, high + offset])

    return value


def _generate_one_row(
    fields: list[EntityField],
    fk_pools: dict[str, list[Any]],
    seen_per_field: dict[str, set],
    unique_fk_queues: dict[str, list[Any]],
    workflows: dict[str, Workflow],
    trends: dict[str, Trend],
    trend_state: dict[str, dict],
    error_injections: dict[str, ErrorInjection],
    geo_routes: dict[str, GeoRoute],
    relationship_lookup: dict[str, dict[Any, dict[str, Any]]],
    relationship_entity_name: dict[str, str],
    previous_row: dict[str, Any] | None,
    position: int,
    count: int,
    injected_fields: set[str] | None = None,
    diagnostics: "GenerationDiagnostics | None" = None,
) -> tuple[dict[str, Any], dict[str, dict[str, Any]]]:
    row: dict[str, Any] = {}
    cross_entity_context: dict[str, dict[str, Any]] = {}

    # Relationship-sourced FK fields are resolved up front, before the main
    # field loop below, so the linked target row is available to every
    # field's formula/rule regardless of declared field order — a formula
    # earlier in `order` than its entity's FK field still needs to see
    # `Customer.age`. A field with its own `formula` keeps formula's
    # existing higher priority (matches the elif ordering below) rather
    # than being silently overridden by its relationship value.
    for field in fields:
        if field.formula or field.name not in relationship_lookup:
            continue
        if field.unique:
            queue = unique_fk_queues[field.name]
            if not queue:
                raise ValueError(
                    f"Field '{field.name}' is unique and ran out of distinct values "
                    "in its pool (relationship or lookup table)"
                )
            value = queue.pop()
        else:
            value = random.choice(fk_pools[field.name])
        row[field.name] = value
        target_row = relationship_lookup[field.name].get(value)
        if target_row is not None:
            cross_entity_context[relationship_entity_name[field.name]] = target_row

    for field in fields:
        workflow_path: list[str] | None = None

        if field.name in row:
            value = row[field.name]
        elif field.formula:
            value = _evaluate_formula(field, {**row, **cross_entity_context})
        elif field.name in trends:
            raw = generate_trend_value(trends[field.name], position, trend_state[field.name])
            value = round(raw) if field.field_type == FieldType.INTEGER else round(raw, 2)
        elif field.name in workflows:
            workflow_path = _generate_state_walk(workflows[field.name])
            value = workflow_path[-1]
        elif field.name in geo_routes:
            route = geo_routes[field.name]
            value = generate_geo_point(
                route.lookup_table.data, route.lat_column, route.lon_column, position, count
            )
        elif not field.required and field.nullable and random.random() < NULLABLE_PROBABILITY:
            value = None
        elif field.name in fk_pools:
            if field.unique:
                queue = unique_fk_queues[field.name]
                if not queue:
                    raise ValueError(
                        f"Field '{field.name}' is unique and ran out of distinct values "
                        "in its pool (relationship or lookup table)"
                    )
                value = queue.pop()
            else:
                value = random.choice(fk_pools[field.name])
        elif field.unique:
            value = _generate_unique_value(field, seen_per_field[field.name], diagnostics)
        else:
            value = _generate_value(field)

        injection = error_injections.get(field.name)
        if injection is not None and random.random() < injection.rate:
            value = _corrupt_value(value, field, injection.error_types, previous_row)
            # Recorded per *candidate*, not per row. The caller only counts
            # it as surviving if this candidate goes on to pass the rules —
            # which is the whole point, since corruption runs before rule
            # checking and a rule on the same field silently undoes it.
            if injected_fields is not None:
                injected_fields.add(field.name)

        row[field.name] = value
        if workflow_path is not None:
            row[f"{field.name}_history"] = workflow_path
    return row, cross_entity_context


def iter_rows(
    fields: list[EntityField],
    count: int,
    fk_pools: dict[str, list[Any]] | None = None,
    rules: list[Rule] | None = None,
    workflows: list[Workflow] | None = None,
    trends: list[Trend] | None = None,
    error_injections: list[ErrorInjection] | None = None,
    event_triggers: list[EventTrigger] | None = None,
    geo_routes: list[GeoRoute] | None = None,
    relationship_lookup: dict[str, dict[Any, dict[str, Any]]] | None = None,
    relationship_entity_name: dict[str, str] | None = None,
    diagnostics: GenerationDiagnostics | None = None,
    start_position: int = 0,
    trend_state: dict[str, dict] | None = None,
) -> Iterator[dict[str, Any]]:
    """Yield `count` rows for `fields`, one at a time.

    Streaming rather than list-building is what lets a job write far more
    rows than fit in memory (see app.services.jobs): only the previous row
    is retained, because that's all the generation loop ever looked at.
    `generate_rows` below wraps this for every caller that genuinely wants
    the whole batch at once, so nothing about existing behaviour changes.

    Note this still takes `count` up front — trends and geo routes are
    functions of a row's position *within the batch*, so the total has to
    be known before the first row. Streaming saves the output, not the
    parameter.

    `fk_pools` maps a field name to a pool of real values to draw from instead
    of randomizing that field independently — this is how both relationships
    and lookup tables are enforced at generation time (see
    build_lookup_pools; a relationship's caller builds this dict from another
    entity's already-generated column the same way). A pool field marked
    `unique` is drawn without replacement; otherwise values are drawn with
    replacement.

    `rules` are boolean expressions a finished row must satisfy; a row that
    fails one is discarded and regenerated (bounded retries). A discarded
    row's values are not "returned" to unique pools/seen-sets, so rules that
    reject a lot of candidates can exhaust a small unique pool faster than the
    requested `count` would otherwise need — a known tradeoff, not a bug.

    `workflows` are state machines attached to a field; that field's value
    comes from a random walk over the graph (see _generate_state_walk)
    instead of the type-based generator, and the walk itself is exposed
    alongside it as `<field>_history`.

    `trends` make a numeric field's value a function of its row's 0-indexed
    position within this call's batch (see app.services.trends) instead of
    an independent random draw — position always starts at 0 here, so a
    trend replays from its start every `generate_rows` call rather than
    continuing across calls (see Trend's docstring for what that means for a
    WebSocket stream's repeated ticks).

    `error_injections` deliberately corrupt a field's value on some fraction
    of rows, after that value is otherwise fully computed (see
    app.models.error_injection and _corrupt_value). Because corruption runs
    before rule-checking, a rule constraining the same field can discard and
    regenerate the very rows error injection was meant to produce — see the
    ErrorInjection model docstring for that tradeoff.

    `event_triggers` are boolean expressions evaluated against each *kept*
    row (after it has already passed every rule); every trigger that matches
    has its `label` appended to that row's `_triggered_events` list, added
    only when at least one trigger is configured (see EventTrigger's
    docstring — this doesn't discard the row or send anything externally).

    `geo_routes` make an object/json field's value a `{"lat", "lon"}` point
    interpolated along an uploaded waypoint sequence, as a function of the
    row's position within this call's batch (see app.services.geo_routes)
    — the same "function of batch position" idea as `trends`, just for a 2D
    path instead of a scalar curve.

    `relationship_lookup` and `relationship_entity_name` (built by the
    caller — see generate_project) are what let a formula/rule/event
    trigger reference *another entity's* field via `TargetEntity.field`
    syntax (see app.services.expressions), for a target entity connected by
    a Relationship: `relationship_lookup[source_field_name]` maps that
    field's possible values to the *specific* already-generated target row
    that value came from, and `relationship_entity_name[source_field_name]`
    is the name that row gets exposed under. Resolved once per row, before
    any of that row's other fields are computed, so it doesn't matter
    whether the referencing formula's `order` comes before or after the
    relationship field's own. Only available from `generate_project`
    (project-wide generation) — a single-entity `generate` call has no
    other entity's data to draw from, so a cross-entity reference there
    fails with a clear "Unknown variable" error rather than silently
    resolving to nothing.
    """
    fk_pools = fk_pools or {}
    rules = rules or []
    event_triggers = event_triggers or []
    workflows_by_field = {w.field.name: w for w in (workflows or [])}
    trends_by_field = {t.field.name: t for t in (trends or [])}
    # Seeded from the caller when one is passed, and mutated in place, so a
    # `random_walk` continues from where the last call left it rather than
    # springing back to `start`. Callers that pass nothing get the old
    # behaviour exactly: a fresh dict per call.
    if trend_state is None:
        trend_state = {}
    for name in trends_by_field:
        trend_state.setdefault(name, {})
    error_injections_by_field = {ei.field.name: ei for ei in (error_injections or [])}
    geo_routes_by_field = {g.field.name: g for g in (geo_routes or [])}
    relationship_lookup = relationship_lookup or {}
    relationship_entity_name = relationship_entity_name or {}
    seen_per_field: dict[str, set] = {
        f.name: set() for f in fields if f.unique and f.name not in fk_pools
    }
    unique_fk_queues: dict[str, list[Any]] = {}
    for field in fields:
        if field.name in fk_pools and field.unique:
            queue = list(fk_pools[field.name])
            random.shuffle(queue)
            unique_fk_queues[field.name] = queue

    if diagnostics is not None:
        diagnostics.rows_requested = count

    previous_row: dict[str, Any] | None = None
    injected: set[str] = set()
    for offset in range(count):
        position = start_position + offset
        row = None
        cross_entity_context: dict[str, dict[str, Any]] = {}
        for _attempt in range(MAX_RULE_ATTEMPTS if rules else 1):
            injected.clear()
            candidate, candidate_context = _generate_one_row(
                fields,
                fk_pools,
                seen_per_field,
                unique_fk_queues,
                workflows_by_field,
                trends_by_field,
                trend_state,
                error_injections_by_field,
                geo_routes_by_field,
                relationship_lookup,
                relationship_entity_name,
                previous_row,
                position,
                count,
                injected_fields=injected if diagnostics is not None else None,
                diagnostics=diagnostics,
            )
            if diagnostics is not None:
                diagnostics.candidate(injected)
            failing = _first_failing_rule(candidate, rules, candidate_context)
            if failing is None:
                if diagnostics is not None:
                    diagnostics.accepted(injected)
                row = candidate
                cross_entity_context = candidate_context
                break
            if diagnostics is not None:
                diagnostics.rule_rejected(failing.condition)
        if row is None:
            raise ValueError(
                f"Could not generate a row satisfying all rules after {MAX_RULE_ATTEMPTS} "
                "attempts — the rules may be too strict or contradictory"
            )
        if event_triggers:
            row["_triggered_events"] = _evaluate_event_triggers(
                row, event_triggers, cross_entity_context
            )
        previous_row = row
        yield row


def generate_rows(
    fields: list[EntityField],
    count: int,
    fk_pools: dict[str, list[Any]] | None = None,
    rules: list[Rule] | None = None,
    workflows: list[Workflow] | None = None,
    trends: list[Trend] | None = None,
    error_injections: list[ErrorInjection] | None = None,
    event_triggers: list[EventTrigger] | None = None,
    geo_routes: list[GeoRoute] | None = None,
    relationship_lookup: dict[str, dict[Any, dict[str, Any]]] | None = None,
    relationship_entity_name: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """The whole batch at once — see `iter_rows` for the streaming form.

    Kept as the default because almost every caller (an API response, a
    CSV download, one WebSocket tick) genuinely wants a list, and because
    it means the streaming refactor changed no existing behaviour."""
    return list(
        iter_rows(
            fields,
            count,
            fk_pools=fk_pools,
            rules=rules,
            workflows=workflows,
            trends=trends,
            error_injections=error_injections,
            event_triggers=event_triggers,
            geo_routes=geo_routes,
            relationship_lookup=relationship_lookup,
            relationship_entity_name=relationship_entity_name,
        )
    )


def _topological_order(graph: dict[uuid.UUID, Iterable[uuid.UUID]]) -> list[uuid.UUID]:
    """Postorder DFS topological sort: `graph[node]` is the set of nodes that must
    come before `node`. Raises ValueError on a cycle."""
    state: dict[uuid.UUID, str] = {}
    order: list[uuid.UUID] = []

    def visit(node: uuid.UUID) -> None:
        if state.get(node) == "done":
            return
        if state.get(node) == "visiting":
            raise ValueError("Circular relationship detected between entities")
        state[node] = "visiting"
        for dep in graph.get(node, ()):
            visit(dep)
        state[node] = "done"
        order.append(node)

    for node in graph:
        visit(node)

    return order


def generate_project(
    entities: list[Entity],
    relationships: list[Relationship],
    counts: dict[uuid.UUID, int],
) -> dict[uuid.UUID, list[dict[str, Any]]]:
    """Generate every entity in a project, honoring relationships: a source
    entity's foreign-key field is populated from its target entity's
    already-generated rows, so targets are always generated before their
    sources (dependency order via topological sort over the relationship
    graph).

    Each entity's lookup attachments are merged into that same fk_pools dict
    (see build_lookup_pools) after the relationship pools are built, so a
    lookup pool wins if a field somehow ends up targeted by both — not
    cross-validated against each other, same as Trend/Workflow aren't.

    Alongside each relationship's fk_pools entry, also builds the
    `relationship_lookup`/`relationship_entity_name` structures
    `generate_rows` needs for cross-entity formula/rule/event-trigger
    references (`TargetEntity.field` — see that function's docstring and
    app.services.expressions) — a value-to-target-row map keyed by the
    source field's name, and the entity name that row should be exposed
    under. This only works here, not from single-entity generation, since
    it needs another entity's rows to already exist."""
    entities_by_id = {e.id: e for e in entities}
    fields_by_id = {f.id: f for e in entities for f in e.fields}

    graph: dict[uuid.UUID, set[uuid.UUID]] = {e.id: set() for e in entities}
    relationships_by_source: dict[uuid.UUID, list[Relationship]] = {}
    for rel in relationships:
        graph.setdefault(rel.source_entity_id, set()).add(rel.target_entity_id)
        relationships_by_source.setdefault(rel.source_entity_id, []).append(rel)

    order = _topological_order(graph)

    generated: dict[uuid.UUID, list[dict[str, Any]]] = {}
    for entity_id in order:
        entity = entities_by_id.get(entity_id)
        if entity is None or not entity.fields:
            generated[entity_id] = []
            continue

        fk_pools: dict[str, list[Any]] = {}
        relationship_lookup: dict[str, dict[Any, dict[str, Any]]] = {}
        relationship_entity_name: dict[str, str] = {}
        for rel in relationships_by_source.get(entity_id, []):
            if rel.relationship_type == RelationshipType.MANY_TO_MANY:
                # No foreign key on either side — that is what makes it
                # many-to-many. The link is emitted as a join table by
                # `generate_join_tables`, so the source's field is its own
                # key and is generated normally.
                continue
            target_field = fields_by_id[rel.target_field_id]
            source_field = fields_by_id[rel.source_field_id]
            target_entity = entities_by_id[rel.target_entity_id]
            target_rows = generated.get(rel.target_entity_id, [])
            pool = [
                row[target_field.name]
                for row in target_rows
                if row.get(target_field.name) is not None
            ]
            if not pool:
                raise ValueError(
                    f"Cannot generate '{entity.name}.{source_field.name}': "
                    f"'{target_entity.name}.{target_field.name}' has no generated values "
                    "to reference (check that entity has fields and a non-zero count)"
                )
            fk_pools[source_field.name] = pool
            relationship_lookup[source_field.name] = {
                row[target_field.name]: row
                for row in target_rows
                if row.get(target_field.name) is not None
            }
            relationship_entity_name[source_field.name] = target_entity.name

        fk_pools.update(build_lookup_pools(entity.lookup_attachments))

        generated[entity_id] = generate_rows(
            entity.fields,
            counts.get(entity_id, 10),
            fk_pools,
            entity.rules,
            entity.workflows,
            entity.trends,
            entity.error_injections,
            entity.event_triggers,
            entity.geo_routes,
            relationship_lookup,
            relationship_entity_name,
        )

    return generated


def join_table_name(source: Entity, target: Entity) -> str:
    """The conventional name for a link table: both sides, lower-cased,
    joined by an underscore. `student_course`, not `StudentCourse` — every
    warehouse this feeds writes it that way."""
    return f"{source.name.lower()}_{target.name.lower()}"


def generate_join_tables(
    entities: list[Entity],
    relationships: list[Relationship],
    generated: dict[uuid.UUID, list[dict[str, Any]]],
) -> dict[str, list[dict[str, Any]]]:
    """Link rows for every `many_to_many` relationship in the project.

    Returned separately from `generate_project` rather than folded into it,
    because a join table is not an entity: it has no id, no fields, no rules
    and nothing to configure. Squeezing it into a dict keyed by entity id
    would have meant inventing an id for something the schema does not
    contain, and every existing caller of `generate_project` keeps working
    untouched.

    Each source row links to between `min_links` and `max_links` **distinct**
    targets — distinct because a join table with a duplicated pair is a bug
    in every schema that has a unique constraint on it, which is most of
    them. A relationship asking for more links than there are targets is
    capped at the number that exist rather than raising: generating fewer
    links than requested is a smaller surprise than a project that will not
    generate at all because one entity's count is low.
    """
    entities_by_id = {e.id: e for e in entities}
    fields_by_id = {f.id: f for e in entities for f in e.fields}

    tables: dict[str, list[dict[str, Any]]] = {}
    for rel in relationships:
        if rel.relationship_type != RelationshipType.MANY_TO_MANY:
            continue
        source = entities_by_id.get(rel.source_entity_id)
        target = entities_by_id.get(rel.target_entity_id)
        if source is None or target is None:
            continue
        source_field = fields_by_id.get(rel.source_field_id)
        target_field = fields_by_id.get(rel.target_field_id)
        if source_field is None or target_field is None:
            continue

        target_keys = [
            row[target_field.name]
            for row in generated.get(rel.target_entity_id, [])
            if row.get(target_field.name) is not None
        ]
        if not target_keys:
            continue

        low = max(rel.min_links, 0)
        high = max(rel.max_links, low)
        rows: list[dict[str, Any]] = []
        for source_row in generated.get(rel.source_entity_id, []):
            source_key = source_row.get(source_field.name)
            if source_key is None:
                continue
            wanted = min(random.randint(low, high), len(target_keys))
            for target_key in random.sample(target_keys, wanted):
                rows.append({source_field.name: source_key, target_field.name: target_key})

        tables[join_table_name(source, target)] = rows

    return tables


def rows_to_csv(fields: list[EntityField], rows: list[dict[str, Any]]) -> str:
    """Renders the declared fields as CSV columns. Workflow `<field>_history`
    values are intentionally dropped here — they're variable-length arrays,
    not a tabular column — and remain available in the JSON response."""
    fieldnames = [f.name for f in fields]
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                name: json.dumps(value) if isinstance(value, (list, dict)) else value
                for name, value in row.items()
            }
        )
    return buffer.getvalue()


def _cell_value(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value)
    return value


def _sheet_columns(fields: list[EntityField], rows: list[dict[str, Any]]) -> list[str]:
    declared = [f.name for f in fields]
    if not rows:
        return declared
    extra = [k for k in rows[0] if k not in declared]
    return declared + extra


def _write_sheet(ws: Worksheet, fields: list[EntityField], rows: list[dict[str, Any]]) -> None:
    columns = _sheet_columns(fields, rows)
    ws.append(columns)
    for row in rows:
        ws.append([_cell_value(row.get(col)) for col in columns])


_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _safe_sheet_name(name: str) -> str:
    return _INVALID_SHEET_CHARS.sub("_", name)[:31] or "Sheet"


def rows_to_excel(fields: list[EntityField], rows: list[dict[str, Any]]) -> bytes:
    """Unlike CSV, this includes any extra keys generation adds (e.g. a
    workflow field's `<field>_history`) alongside the declared fields —
    Excel isn't a strict fixed-column format the way CSV is."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    _write_sheet(ws, fields, rows)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def project_rows_to_excel(
    entities: list[Entity], generated: dict[uuid.UUID, list[dict[str, Any]]]
) -> bytes:
    """One workbook, one sheet per entity."""
    wb = Workbook()
    wb.remove(wb.active)
    for entity in entities:
        ws = wb.create_sheet(title=_safe_sheet_name(entity.name))
        _write_sheet(ws, entity.fields, generated.get(entity.id, []))
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def project_rows_to_csv_zip(
    entities: list[Entity],
    generated: dict[uuid.UUID, list[dict[str, Any]]],
    join_tables: dict[str, list[dict[str, Any]]] | None = None,
) -> bytes:
    """One `<entity>.csv` file per entity, zipped together — CSV has no
    multi-table concept, so a project-wide export is a zip of flat files.

    Many-to-many join tables land alongside them as their own files. They
    have no `EntityField` list to take column order from, so the columns
    come from the first row's keys — which is exactly the two key names, in
    the order `generate_join_tables` wrote them."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for entity in entities:
            csv_text = rows_to_csv(entity.fields, generated.get(entity.id, []))
            zf.writestr(f"{entity.name}.csv", csv_text)
        for name, rows in (join_tables or {}).items():
            zf.writestr(f"{name}.csv", _plain_rows_to_csv(rows))
    return buffer.getvalue()


def _plain_rows_to_csv(rows: list[dict[str, Any]]) -> str:
    """CSV for rows with no field definitions behind them."""
    if not rows:
        return ""
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue()
